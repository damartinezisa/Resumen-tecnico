#!/usr/bin/env python3
"""
Script to extract tables from PDF to Excel without using OCR.
This is the first step in a two-step process:
1. Extract PDF tables to Excel for pre-validation (this script)
2. Convert validated Excel to JSON (use excel-to-json.py)

This approach ensures no OCR errors with numbers and allows human validation.
"""

# Import required libraries for PDF processing, Excel creation, and file handling
import pdfplumber  # Library for extracting tables directly from PDF structure
import pandas as pd  # Library for data manipulation and Excel file creation
import sys  # System-specific parameters and functions (for command-line arguments)
import os  # Operating system interface (for file path operations)
import openpyxl  # Library for reading and writing Excel files with formatting
from openpyxl.styles import PatternFill, Font, Alignment  # Excel cell styling


def extract_pdf_to_excel(pdf_path: str, page_number: int = 2, output_excel: str = None):
    """
    Extract tables from a specific PDF page and save to Excel for validation.
    
    This function uses pdfplumber to detect and extract table structures directly
    from the PDF without OCR. The extracted data is saved to an Excel file with
    multiple sheets for easy validation and correction.
    
    Args:
        pdf_path: Path to the PDF file to process
        page_number: Page number to extract (1-indexed, default is 2)
        output_excel: Output Excel file path (default: auto-generated from PDF name)
    
    Returns:
        Path to the created Excel file
    
    Raises:
        ValueError: If the specified page number exceeds the total pages in PDF
    """
    # If no output filename is specified, generate one based on the PDF name
    if output_excel is None:
        base_name = os.path.splitext(pdf_path)[0]  # Remove .pdf extension
        output_excel = f"{base_name}_page{page_number}_tables.xlsx"
    
    # Print progress message to stderr (not stdout, so it doesn't interfere with piped output)
    print(f"Extracting tables from {pdf_path}, page {page_number}...", file=sys.stderr)
    
    # Open and process the PDF file
    with pdfplumber.open(pdf_path) as pdf:
        # Validate that the requested page exists in the PDF
        if page_number > len(pdf.pages):
            raise ValueError(f"PDF only has {len(pdf.pages)} pages, cannot extract page {page_number}")
        
        # Get the specified page (convert from 1-indexed to 0-indexed)
        page = pdf.pages[page_number - 1]
        
        # Extract tables using line-based detection
        # This uses the PDF's internal structure to detect table boundaries
        tables = page.extract_tables({
            "vertical_strategy": "lines",  # Detect vertical table lines from PDF structure
            "horizontal_strategy": "lines",  # Detect horizontal table lines from PDF structure
            "snap_tolerance": 3,  # Pixels tolerance for aligning nearby lines
            "join_tolerance": 3,  # Pixels tolerance for joining line segments
        })
        
        # Report how many tables were detected
        print(f"Found {len(tables)} table(s)", file=sys.stderr)
        
        # Create an Excel writer object to save multiple sheets
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # Handle case where no tables are found
            if not tables:
                print("Warning: No tables found!", file=sys.stderr)
                # Create an information sheet with a warning message
                df = pd.DataFrame(["No tables found on this page"])
                df.to_excel(writer, sheet_name='Info', index=False, header=False)
            else:
                # Process the main table (which contains all 3 logical tables merged)
                main_table = tables[0]
                
                # Create a DataFrame from the raw table data
                # This preserves the original extracted structure for reference
                df = pd.DataFrame(main_table)
                
                # Save raw extraction to first sheet for debugging/reference
                df.to_excel(writer, sheet_name='Raw_Data', index=False, header=False)
                
                # Now parse and separate the three logical tables within the main table
                # The PDF contains 3 tables that appear as one continuous table:
                # 1. ANALISIS table (rows 0-32 approximately)
                # 2. PRODUCTO TERMINADO (AZUCAR) table (rows 33-37 approximately)
                # 3. PRODUCTO TERMINADO (AZUCAR) - CONTINUACIÓN table (rows 38+ approximately)
                
                # Initialize lists to store data for each logical table
                analisis_data = []
                producto_data = []
                continuacion_data = []
                
                # Variables to track where each section starts
                producto_start = None
                continuacion_start = None
                
                # Scan through all rows to find section boundaries
                for i, row in enumerate(main_table):
                    # Get the first column value and clean it for comparison
                    first_col = str(row[0]) if row[0] else ""
                    first_col_clean = first_col.replace("*", "").replace(" ", "").upper()
                    
                    # Look for "PRODUCTO TERMINADO" keyword to identify second section
                    if "PRODUCTO" in first_col_clean and "TERMINADO" in first_col_clean:
                        producto_start = i
                    # Look for "DESCRIPCION" keyword after row 30 with "Quintales" in the row
                    # This identifies the start of the third section (CONTINUACIÓN)
                    elif "DESCRIPCION" in first_col.upper() and i > 30 and "Quintales" in " ".join([str(c) for c in row if c]):
                        continuacion_start = i
                        break  # Found the last section, no need to continue scanning
                
                # Split the main table data into the three logical sections
                if producto_start and continuacion_start:
                    # All three sections were found
                    analisis_data = main_table[:producto_start]
                    producto_data = main_table[producto_start:continuacion_start]
                    continuacion_data = main_table[continuacion_start:]
                elif producto_start:
                    # Only found first two sections
                    analisis_data = main_table[:producto_start]
                    producto_data = main_table[producto_start:]
                else:
                    # Only found ANALISIS section (fallback)
                    analisis_data = main_table
                
                # Save each separated table to its own Excel sheet
                # This makes it easier to validate and correct each table independently
                
                if analisis_data:
                    df_analisis = pd.DataFrame(analisis_data)
                    df_analisis.to_excel(writer, sheet_name='ANALISIS', index=False, header=False)
                
                if producto_data:
                    df_producto = pd.DataFrame(producto_data)
                    df_producto.to_excel(writer, sheet_name='PRODUCTO_TERMINADO', index=False, header=False)
                
                if continuacion_data:
                    df_continuacion = pd.DataFrame(continuacion_data)
                    df_continuacion.to_excel(writer, sheet_name='CONTINUACION', index=False, header=False)
        
        # Apply formatting to make the Excel file easier to read and validate
        # This step enhances the user experience when reviewing the extracted data
        wb = openpyxl.load_workbook(output_excel)
        
        # Format each sheet in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Auto-adjust column widths based on content
            # This makes all content visible without manual column resizing
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Find the maximum content length in this column
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Set column width (add padding, but cap at 50 characters)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Highlight header rows with yellow background for easy identification
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Check first 3 rows for header keywords and apply formatting
            for row in ws.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    # If cell contains common header keywords, highlight it
                    if cell.value and any(keyword in str(cell.value).upper() 
                                         for keyword in ['ANALISIS', 'HOY', 'HASTA', 'BRIX', 
                                                        'PRODUCTO', 'DESCRIPCION']):
                        cell.fill = yellow_fill  # Yellow background
                        cell.font = Font(bold=True)  # Bold text
        
        # Save the formatted workbook
        wb.save(output_excel)
        
    # Print success messages with next steps
    print(f"✓ Excel file created: {output_excel}", file=sys.stderr)
    print(f"✓ Please validate the data in Excel and save any corrections.", file=sys.stderr)
    print(f"✓ Then use 'python excel-to-json.py {output_excel}' to convert to JSON", file=sys.stderr)
    
    return output_excel


def main():
    """
    Main function to handle command-line execution.
    
    Usage:
        python pdf-to-excel.py <pdf_file> [page_number]
    
    Examples:
        python pdf-to-excel.py 1003.pdf 2
        python pdf-to-excel.py report.pdf
    """
    # Check if required arguments are provided
    if len(sys.argv) < 2:
        # Print usage instructions if no arguments provided
        print("Usage:")
        print(f"    python {sys.argv[0]} <pdf_file> [page_number]")
        print(f"")
        print(f"Arguments:")
        print(f"    pdf_file      - Path to the PDF file to process (required)")
        print(f"    page_number   - Page number to extract (default: 2)")
        print(f"")
        print(f"Example:")
        print(f"    python {sys.argv[0]} 1003.pdf 2")
        print(f"")
        print(f"Output:")
        print(f"    Creates an Excel file with extracted tables for validation")
        print(f"    Then use excel-to-json.py to convert the validated data to JSON")
        sys.exit(1)
    
    # Get command-line arguments
    pdf_path = sys.argv[1]
    page_number = int(sys.argv[2]) if len(sys.argv) > 2 else 2  # Default to page 2
    
    # Execute the extraction
    excel_path = extract_pdf_to_excel(pdf_path, page_number)


# Standard Python idiom to execute main() when script is run directly
if __name__ == "__main__":
    main()
