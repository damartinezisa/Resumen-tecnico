#!/usr/bin/env python3
"""
Script to extract tables from PDF to Excel without using OCR.
This provides a two-step process:
1. Extract PDF tables to Excel for pre-validation
2. Convert validated Excel to JSON
"""

import pdfplumber
import pandas as pd
import json
import sys
import os
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


def extract_pdf_to_excel(pdf_path: str, page_number: int = 2, output_excel: str = None):
    """
    Extract tables from PDF page and save to Excel for validation.
    
    Args:
        pdf_path: Path to the PDF file
        page_number: Page number to extract (1-indexed)
        output_excel: Output Excel file path (default: based on PDF name)
    
    Returns:
        Path to the created Excel file
    """
    if output_excel is None:
        base_name = os.path.splitext(pdf_path)[0]
        output_excel = f"{base_name}_page{page_number}_tables.xlsx"
    
    print(f"Extracting tables from {pdf_path}, page {page_number}...", file=sys.stderr)
    
    with pdfplumber.open(pdf_path) as pdf:
        if page_number > len(pdf.pages):
            raise ValueError(f"PDF only has {len(pdf.pages)} pages, cannot extract page {page_number}")
        
        # Get the specified page (1-indexed)
        page = pdf.pages[page_number - 1]
        
        # Extract tables using line-based detection
        tables = page.extract_tables({
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 3,
            "join_tolerance": 3,
        })
        
        print(f"Found {len(tables)} table(s)", file=sys.stderr)
        
        # Create Excel writer
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            if not tables:
                print("Warning: No tables found!", file=sys.stderr)
                # Create empty sheet
                df = pd.DataFrame(["No tables found on this page"])
                df.to_excel(writer, sheet_name='Info', index=False, header=False)
            else:
                # Process the main table (should contain all 3 tables merged)
                main_table = tables[0]
                
                # Create a DataFrame from the raw table
                df = pd.DataFrame(main_table)
                
                # Save raw extraction to first sheet
                df.to_excel(writer, sheet_name='Raw_Data', index=False, header=False)
                
                # Now parse and separate the three logical tables
                # Based on the structure, we need to identify:
                # 1. ANALISIS table
                # 2. PRODUCTO TERMINADO (AZUCAR) table
                # 3. PRODUCTO TERMINADO (AZUCAR) - CONTINUACIÓN table
                
                analisis_data = []
                producto_data = []
                continuacion_data = []
                
                current_section = None
                
                for i, row in enumerate(main_table):
                    # Identify sections by keywords in first column
                    first_col = str(row[0]) if row[0] else ""
                    
                    if "ANALISIS" in first_col.upper() and i < 5:
                        current_section = "ANALISIS"
                        analisis_data.append(row)
                    elif "PRODUCTO TERMINADO" in first_col.upper() and "CONTINUACIÓN" not in first_col.upper() and "CONTINUACION" not in first_col.upper():
                        current_section = "PRODUCTO"
                        producto_data.append(row)
                    elif "CONTINUACIÓN" in first_col.upper() or "CONTINUACION" in first_col.upper():
                        current_section = "CONTINUACION"
                        continuacion_data.append(row)
                    else:
                        # Add row to current section
                        if current_section == "ANALISIS":
                            analisis_data.append(row)
                        elif current_section == "PRODUCTO":
                            producto_data.append(row)
                        elif current_section == "CONTINUACION":
                            continuacion_data.append(row)
                
                # Save separated tables to individual sheets
                if analisis_data:
                    df_analisis = pd.DataFrame(analisis_data)
                    df_analisis.to_excel(writer, sheet_name='ANALISIS', index=False, header=False)
                
                if producto_data:
                    df_producto = pd.DataFrame(producto_data)
                    df_producto.to_excel(writer, sheet_name='PRODUCTO_TERMINADO', index=False, header=False)
                
                if continuacion_data:
                    df_continuacion = pd.DataFrame(continuacion_data)
                    df_continuacion.to_excel(writer, sheet_name='CONTINUACION', index=False, header=False)
        
        # Apply formatting to make validation easier
        wb = openpyxl.load_workbook(output_excel)
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Highlight header rows (yellow background)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in ws.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    if cell.value and any(keyword in str(cell.value).upper() for keyword in ['ANALISIS', 'HOY', 'HASTA', 'BRIX', 'PRODUCTO', 'DESCRIPCION']):
                        cell.fill = yellow_fill
                        cell.font = Font(bold=True)
        
        wb.save(output_excel)
        
    print(f"✓ Excel file created: {output_excel}", file=sys.stderr)
    print(f"✓ Please validate the data in Excel and save any corrections.", file=sys.stderr)
    print(f"✓ Then use 'python {sys.argv[0]} --excel-to-json {output_excel}' to convert to JSON", file=sys.stderr)
    
    return output_excel


def excel_to_json(excel_path: str, output_json: str = None):
    """
    Convert validated Excel file to JSON format.
    
    Args:
        excel_path: Path to the Excel file
        output_json: Output JSON file path (default: based on Excel name)
    
    Returns:
        Path to the created JSON file
    """
    if output_json is None:
        base_name = os.path.splitext(excel_path)[0]
        output_json = f"{base_name}.json"
    
    print(f"Converting Excel to JSON: {excel_path}...", file=sys.stderr)
    
    # Read Excel file
    xl_file = pd.ExcelFile(excel_path)
    
    results = []
    
    # Check which sheets exist
    has_analisis = 'ANALISIS' in xl_file.sheet_names
    has_producto = 'PRODUCTO_TERMINADO' in xl_file.sheet_names
    has_continuacion = 'CONTINUACION' in xl_file.sheet_names
    
    # Parse ANALISIS sheet
    if has_analisis:
        df_analisis = pd.read_excel(excel_path, sheet_name='ANALISIS', header=None)
        hoy_indicators, hasta_indicators = parse_analisis_from_df(df_analisis)
        
        results.append({
            "Nombre Agrupador": "ANALISIS - HOY",
            "Indicadores": hoy_indicators
        })
        results.append({
            "Nombre Agrupador": "ANALISIS - HASTA",
            "Indicadores": hasta_indicators
        })
    
    # Parse PRODUCTO TERMINADO sheet
    if has_producto:
        df_producto = pd.read_excel(excel_path, sheet_name='PRODUCTO_TERMINADO', header=None)
        producto_indicators = parse_producto_from_df(df_producto)
        
        results.append({
            "Nombre Agrupador": "PRODUCTO TERMINADO (AZUCAR)",
            "Indicadores": producto_indicators
        })
    
    # Parse CONTINUACION sheet
    if has_continuacion:
        df_continuacion = pd.read_excel(excel_path, sheet_name='CONTINUACION', header=None)
        continuacion_indicators = parse_continuacion_from_df(df_continuacion)
        
        results.append({
            "Nombre Agrupador": "PRODUCTO TERMINADO (AZUCAR) - CONTINUACIÓN",
            "Indicadores": continuacion_indicators
        })
    
    # Write JSON
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4, ensure_ascii=False)
    
    print(f"✓ JSON file created: {output_json}", file=sys.stderr)
    
    return output_json


def parse_analisis_from_df(df):
    """Parse ANALISIS table from DataFrame"""
    hoy_indicators = []
    hasta_indicators = []
    
    # Find header rows
    # Expected structure: Row 0 has "ANALISIS", "HOY", "HASTA"
    # Row 1 has column names: "BRIX", "SAC", "PZA", "COLOR", "PH", "GR"
    
    # Find which columns correspond to HOY and HASTA
    hoy_cols = {}
    hasta_cols = {}
    
    # Scan first few rows to identify column structure
    for row_idx in range(min(3, len(df))):
        row = df.iloc[row_idx]
        for col_idx, val in enumerate(row):
            val_str = str(val).strip() if pd.notna(val) else ""
            
            # Map columns based on headers
            if "BRIX" in val_str.upper():
                # Determine if this is HOY or HASTA column
                # Look at previous columns for HOY/HASTA indicator
                context = " ".join([str(df.iloc[r, max(0, col_idx-5):col_idx+1].values) for r in range(row_idx+1)])
                if "HOY" in context and "BRIX" not in hoy_cols:
                    hoy_cols["BRIX"] = col_idx
                elif "HASTA" in context and "BRIX" not in hasta_cols:
                    hasta_cols["BRIX"] = col_idx
    
    # For simplicity, use fixed column mapping based on visual inspection
    # HOY: columns around index 3-8 for BRIX, SAC, PZA, then 16-21 for COLOR, PH, GR
    # HASTA: columns around index 9-14 for BRIX, SAC, PZA
    
    # Process data rows (skip header rows)
    for row_idx in range(2, len(df)):
        row = df.iloc[row_idx]
        
        # First column is DESCRIPCION
        descripcion = str(row[0]).strip() if pd.notna(row[0]) else ""
        
        if not descripcion or descripcion == "" or "nan" in descripcion.lower():
            continue
        
        # Skip separator rows
        if all(c in '—-_= \t' for c in descripcion):
            continue
        
        # Create HOY indicator
        hoy_indicator = {
            "DESCRIPCION": descripcion,
            "BRIX": parse_value(row[3]),
            "SAC": parse_value(row[4]),
            "PZA": parse_value(row[7]),
            "COLOR": parse_value(row[16]),
            "PH": parse_value(row[18]),
            "GR": parse_value(row[20])
        }
        
        # Create HASTA indicator
        hasta_indicator = {
            "DESCRIPCION": descripcion,
            "BRIX": parse_value(row[9]),
            "SAC": parse_value(row[11]),
            "PZA": parse_value(row[14])
        }
        
        # Only add if we have at least the description
        if descripcion and len(descripcion) > 1:
            hoy_indicators.append(hoy_indicator)
            hasta_indicators.append(hasta_indicator)
    
    return hoy_indicators, hasta_indicators


def parse_producto_from_df(df):
    """Parse PRODUCTO TERMINADO table from DataFrame"""
    indicators = []
    
    # This table is typically transposed - headers are in columns
    # Find the row with ESTANDAR/DIA, HOY, HASTA
    
    # Look for rows containing these keywords
    estandar_row_idx = None
    hoy_row_idx = None
    hasta_row_idx = None
    
    for row_idx in range(len(df)):
        row = df.iloc[row_idx]
        first_val = str(row[0]).strip().upper() if pd.notna(row[0]) else ""
        
        if "ESTANDAR" in first_val:
            estandar_row_idx = row_idx
        elif first_val == "HOY":
            hoy_row_idx = row_idx
        elif first_val == "HASTA":
            hasta_row_idx = row_idx
    
    # Find header row with column names
    header_row_idx = None
    for row_idx in range(min(5, len(df))):
        row = df.iloc[row_idx]
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        if "TOTAL" in row_str.upper() and "QQ" in row_str.upper() and "CRUDA" in row_str.upper():
            header_row_idx = row_idx
            break
    
    if header_row_idx is not None and estandar_row_idx is not None:
        # Parse column headers
        header_row = df.iloc[header_row_idx]
        
        # Expected columns (combined from multiple cells):
        # TOTAL QQ CRUDA, TOTAL QQ ESTAN., TOTAL QQ REFIN., QQ PRODUCIDOS, 
        # AZ. EQUIV. (MIEL), AZ. PMR, TOTAL QUINTALES
        
        column_descriptions = [
            "TOTAL QQ CRUDA",
            "TOTAL QQ ESTAN.",
            "TOTAL QQ REFIN.",
            "QQ PRODUCIDOS",
            "AZ. EQUIV. (MIEL)",
            "AZ. PMR",
            "TOTAL QUINTALES"
        ]
        
        # For each column description, get values from ESTANDAR, HOY, HASTA rows
        # We need to map column indices - typically starts around column 3
        value_cols = []
        for col_idx in range(len(df.columns)):
            # Check if this column has data in the value rows
            if estandar_row_idx and pd.notna(df.iloc[estandar_row_idx, col_idx]):
                val = str(df.iloc[estandar_row_idx, col_idx]).strip()
                if val and val != "nan" and any(c.isdigit() for c in val):
                    value_cols.append(col_idx)
        
        # Create indicators
        for i, desc in enumerate(column_descriptions):
            indicator = {"DESCRIPCION": desc}
            
            # Get value from appropriate column
            if i < len(value_cols):
                col_idx = value_cols[i]
                indicator["ESTANDAR/DIA"] = parse_value(df.iloc[estandar_row_idx, col_idx])
                indicator["HOY"] = parse_value(df.iloc[hoy_row_idx, col_idx]) if hoy_row_idx else None
                indicator["HASTA"] = parse_value(df.iloc[hasta_row_idx, col_idx]) if hasta_row_idx else None
            else:
                indicator["ESTANDAR/DIA"] = None
                indicator["HOY"] = None
                indicator["HASTA"] = None
            
            indicators.append(indicator)
    
    return indicators


def parse_continuacion_from_df(df):
    """Parse CONTINUACION table from DataFrame"""
    indicators = []
    
    # Find header row
    header_row_idx = None
    for row_idx in range(min(5, len(df))):
        row = df.iloc[row_idx]
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        if "DESCRIPCION" in row_str.upper() and "Quintales" in row_str:
            header_row_idx = row_idx
            break
    
    if header_row_idx is None:
        return indicators
    
    # Column names
    column_names = ["DESCRIPCION", "Quintales", "% HUM.", "% CEN.", "% POL.", "COLOR", 
                    "Mg/kg Vit.\"A\"", "T. GRANO", "% C.V", "FS", "TEMP. ºC.", "SED."]
    
    # Process data rows
    for row_idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[row_idx]
        
        descripcion = str(row[0]).strip() if pd.notna(row[0]) else ""
        
        if not descripcion or descripcion == "nan":
            continue
        
        # Skip separator rows
        if all(c in '—-_= \t' for c in descripcion):
            continue
        
        indicator = {"DESCRIPCION": descripcion}
        
        # Map remaining columns (starting from column 1)
        for i, col_name in enumerate(column_names[1:]):
            col_idx = i + 1
            if col_idx < len(row):
                indicator[col_name] = parse_value(row[col_idx])
            else:
                indicator[col_name] = None
        
        indicators.append(indicator)
    
    return indicators


def parse_value(val):
    """Parse a value from Excel cell to appropriate Python type"""
    if pd.isna(val):
        return None
    
    val_str = str(val).strip()
    
    if not val_str or val_str == "" or val_str.lower() == "nan":
        return None
    
    # Remove thousands separators (commas)
    val_str = val_str.replace(',', '')
    
    # Try to parse as number
    try:
        if '.' in val_str:
            return float(val_str)
        else:
            return int(val_str)
    except ValueError:
        # Return as string if not a number
        return val_str if val_str else None


def main():
    """Main function"""
    if len(sys.argv) < 2:
        print("Usage:")
        print(f"  Step 1 - Extract PDF to Excel:")
        print(f"    python {sys.argv[0]} <pdf_file> [page_number]")
        print(f"")
        print(f"  Step 2 - Convert Excel to JSON:")
        print(f"    python {sys.argv[0]} --excel-to-json <excel_file> [output_json]")
        print(f"")
        print(f"Example:")
        print(f"  python {sys.argv[0]} 1003.pdf 2")
        print(f"  # Validate data in Excel, then:")
        print(f"  python {sys.argv[0]} --excel-to-json 1003_page2_tables.xlsx")
        sys.exit(1)
    
    if sys.argv[1] == "--excel-to-json":
        # Convert Excel to JSON
        if len(sys.argv) < 3:
            print("Error: Excel file path required", file=sys.stderr)
            sys.exit(1)
        
        excel_path = sys.argv[2]
        output_json = sys.argv[3] if len(sys.argv) > 3 else None
        
        json_path = excel_to_json(excel_path, output_json)
        
        # Output only JSON to stdout
        with open(json_path, 'r', encoding='utf-8') as f:
            print(f.read())
    
    else:
        # Extract PDF to Excel
        pdf_path = sys.argv[1]
        page_number = int(sys.argv[2]) if len(sys.argv) > 2 else 2
        
        excel_path = extract_pdf_to_excel(pdf_path, page_number)


if __name__ == "__main__":
    main()
